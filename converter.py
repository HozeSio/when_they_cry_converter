#!/usr/bin/env python3
import sys
import os
import re
import openpyxl
import translation_extractor

method_pattern = re.compile(r"""
(                       # method front part [0]
\s*OutputLine
\s*\(
\s*
)
([^,]*)                    # first parameter (actor) [1]
(\s*,\s*)              # comma [2]
(.*)                  # second parameter (text) [3]
([ \t\x0B\f\r]*,\n?[ \t\x0B\f\r]*) # comma [4]
([^,\n]*)                    # third parameter (actor-alt) [5]
(\s*,\s*)               # comma [6]
(.*)                  # fourth parameter (text-alt) [7]
(\s*,\s*)               # comma [8]
(.*)                    # fifth parameter (show option) [9]
(                       # last part [10]
\s*\)
\s*;
)
""", re.VERBOSE | re.MULTILINE)


def validate_folder(folder_path: str):
    result = True
    for file_name in os.listdir(folder_path):
        if not file_name.endswith('.txt'):
            continue

        print(f"validating {file_name}")
        file_path = os.path.join(folder_path, file_name)
        with open(file_path, 'r', encoding='utf-8') as f:
            result &= validate_text(f.read())

    if not result:
        exit(-1)


def validate_text(text: str):
    for method_match in method_pattern.finditer(text):
        groups = method_match.groups()
        # output text
        if groups[1] == 'NULL':
            if not groups[3].startswith('\"') or not groups[3].endswith('\"') or not groups[7].startswith('\"') or not groups[7].endswith('\"'):
                print(method_match.group(), file=sys.stderr)
                return False
        # set actor
        else:
            pass
    return True


class TextConverter:
    def __init__(self):
        self.match_count = 0
        self.sentences = []
        self.last_actor = None
        self.actor_pattern = re.compile(r'"<color=.*>(.*)</color>"')
        self.remove_quotation_mark_pattern = re.compile(r'"(.*)"')
        self.translation = {}

    def strip_quotation_mark(self, text: str):
        m = self.remove_quotation_mark_pattern.match(text)
        if not m:
            print(text)
            raise Exception
        return m.groups()[0]

    def extract_from_match(self, match_obj):
        groups = match_obj.groups()
        param1 = groups[1]
        param2 = groups[3]
        param3 = groups[5]
        param4 = groups[7]
        param5 = groups[9]

        if param1 != 'NULL':  # actor setting
            self.last_actor = self.actor_pattern.match(param3).groups()[0]
            return

        if param2.startswith('\"<size=') or param4.startswith('\"<size='):
            return

        # store sentence
        self.sentences.append((self.last_actor, self.strip_quotation_mark(param2), self.strip_quotation_mark(param4)))
        self.last_actor = None
        self.match_count += 1

    def extract_text(self, text):
        for match in method_pattern.finditer(text):
            self.extract_from_match(match)

    def repl_replace_text(self, match_obj) -> str:
        groups = list(match_obj.groups())
        param1 = groups[1]
        param2 = groups[3]
        param3 = groups[5]
        param4 = groups[7]
        param5 = groups[9]

        if param1 != 'NULL':  # actor setting
            self.last_actor = self.actor_pattern.match(param3).groups()[0]
            return match_obj.group()

        if param2.startswith('\"<size=') or param4.startswith('\"<size='):
            return match_obj.group()

        # replace english text to translation text based on japanese text
        try:
            key = self.strip_quotation_mark(param2)
            # empty text handling
            if not key:
                key = None
            translated_text = self.translation[key]
        except KeyError:
            print(match_obj.groups())
            raise
        groups[7] = f'\"{translated_text}\"'
        return "".join(groups)

    def replace_text(self, text, translation: {}):
        self.translation = translation
        return method_pattern.sub(self.repl_replace_text, text)


class FolderConverter:
    def __init__(self, folder_path):
        self.folder_path = os.path.normpath(folder_path)
        (self.folder_directory, self.folder_name) = os.path.split(self.folder_path)

    def export_text(self):
        converted_folder = os.path.join(self.folder_directory, self.folder_name + '_converted')
        if not os.path.exists(converted_folder):
            os.mkdir(converted_folder)

        for file_name in os.listdir(self.folder_path):
            if not file_name.endswith('.txt'):
                continue

            file_path = os.path.join(self.folder_path, file_name)
            with open(file_path, 'r', encoding='utf-8') as f:
                print(f"start converting {file_name}....", end='')
                text = f.read()
                if not validate_text(text):
                    exit(-1)
                file_name_only = os.path.splitext(file_name)[0]
                text_converter = TextConverter()
                text_converter.extract_text(text)

                # write xlsx content
                wb = openpyxl.Workbook()
                # wb.remove(wb.active)
                # ws = wb.create_sheet(file_name)
                ws = wb.active
                ws.append(['actor', 'japanese', 'english', 'translation'])
                for sentence in text_converter.sentences:
                    ws.append(sentence)
                wb.save(os.path.join(converted_folder, file_name_only + '.xlsx'))
                wb.close()
                print(f"now converted to {file_name_only}.xlsx")

    def replace_text(self, translation_folder):
        replaced_folder = os.path.join(self.folder_directory, self.folder_name + '_replaced')
        if not os.path.exists(replaced_folder):
            os.mkdir(replaced_folder)

        translation_folder = os.path.normpath(translation_folder)
        for file_name in os.listdir(translation_folder):
            file_name_only = os.path.splitext(file_name)[0]
            script_file_name = f'{file_name_only}.txt'
            script_path = os.path.join(self.folder_path, script_file_name)
            if not file_name.endswith('.xlsx') or not os.path.exists(script_path):
                continue
            print(f'start replacing {file_name_only}....', end='')

            file_path = os.path.join(translation_folder, file_name)
            wb = openpyxl.open(file_path)
            ws = wb.active
            translation = {}
            for row in ws.rows:
                translation[row[2].value] = row[4].value
            wb.close()

            with open(script_path, 'r', encoding='utf-8') as f:
                text_converter = TextConverter()
                replaced_text = text_converter.replace_text(f.read(), translation)
                with open(os.path.join(replaced_folder, script_file_name), 'w', encoding='utf-8') as o:
                    o.write(replaced_text)

            print('finished')


def combine_xlsx(original_folder, translated_folder):
    for file_name in os.listdir(translated_folder):
        if not file_name.endswith('.xlsx'):
            continue
        file_name = file_name.replace('kor', '')

        original_path = os.path.join(original_folder, file_name)
        if not os.path.exists(original_path):
            continue

        original_wb = openpyxl.open(original_path)
        original_ws = original_wb.active

        translated_wb = openpyxl.open(os.path.join(translated_folder, file_name))
        translated_ws = translated_wb.active

        for index, row in enumerate(translated_ws.iter_rows(), 1):
            original_ws.cell(row=index, column=4).value = row[2].value

        original_wb.save(original_path)
        original_wb.close()


def insert_actor_column(old_folder, actor_folder):
    for file_name in os.listdir(old_folder):
        if not file_name.endswith('.xlsx'):
            continue

        old_path = os.path.join(old_folder, file_name)
        old_wb = openpyxl.open(old_path)
        old_ws = old_wb.active

        actor_wb = openpyxl.open(os.path.join(actor_folder, file_name))
        actor_ws = actor_wb.active

        for index, row in enumerate(actor_ws.iter_rows(), 1):
            if old_ws.cell(row=index, column=2).value != row[2].value:
                print(f"{file_name} has different row at {index} {old_ws.cell(row=index, column=2).value} != {row[2].value}")
                break

        old_ws.insert_cols(2)

        for index, row in enumerate(actor_ws.iter_rows(), 1):
            old_ws.cell(row=index, column=2).value = row[1].value

        old_wb.save(old_path)
        old_wb.close()


if __name__ == '__main__':
    if len(sys.argv) == 1 or sys.argv[1] == 'help':
        print(
"""
usage: converter.py [commands]
available commands:
    export_text <Update folder>
    - export text parameter to xlsx file from the script
    replace_text <Update folder> <translation folder>
    - Replace english text to translated text
    extract_text <file_path>
    - extract text line from the onscript file and export to xlsx
    combine_xlsx <original_folder> <translated_folder>
    insert_actor_column <old_folder> <actor_folder>
"""
        )
    elif sys.argv[1] == 'export_text':
        converter = FolderConverter(sys.argv[2])
        converter.export_text()
    elif sys.argv[1] == 'replace_text':
        converter = FolderConverter(sys.argv[2])
        converter.replace_text(sys.argv[3])
    elif sys.argv[1] == 'validate_folder':
        validate_folder(sys.argv[2])
    elif sys.argv[1] == 'extract_text':
        extractor = translation_extractor.TextExtractor()
        extractor.extract_text(sys.argv[2])
    elif sys.argv[1] == 'combine_xlsx':
        combine_xlsx(sys.argv[2], sys.argv[3])
    elif sys.argv[1] == 'insert_actor_column':
        insert_actor_column(sys.argv[2], sys.argv[3])
    else:
        print("invalid command", file=sys.stderr)
        exit(-1)
